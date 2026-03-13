import io
import os
from calendar import monthrange
from datetime import date, datetime

import openpyxl
from fastapi import APIRouter, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from lib.auth import require_auth, get_session
from lib.supabase import supabase

router = APIRouter()
templates = Jinja2Templates(directory="templates")

TEMPLATES_DIR = "static/templates"

MOIS_FR = {
    1: "Janvier", 2: "Février", 3: "Mars", 4: "Avril",
    5: "Mai", 6: "Juin", 7: "Juillet", 8: "Août",
    9: "Septembre", 10: "Octobre", 11: "Novembre", 12: "Décembre",
}

def get_user_email(request: Request) -> str:
    session = get_session(request)
    return session["user"].email if session else ""

# ─── Helpers ──────────────────────────────────────────

def build_libelle(inv: dict) -> str:
    parts = [inv.get("vendor_name", "").strip()]
    if inv.get("detail"):
        parts.append(inv["detail"].strip())
    libelle = " : ".join(filter(None, parts))

    if inv.get("payment_date"):
        try:
            pd_obj = datetime.strptime(inv["payment_date"], "%Y-%m-%d")
            pd_fmt = pd_obj.strftime("%d/%m/%Y")
        except Exception:
            pd_fmt = inv["payment_date"]
        suffix = f"Réglé le {pd_fmt}"
        if inv.get("payment_account"):
            suffix += f" ({inv['payment_account']})"
        libelle += f". {suffix}"

    return libelle


def get_day(inv: dict):
    date_str = inv.get("invoice_date")
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").day
    except Exception:
        return None


def make_piece_number(annee: int, idx: int) -> int:
    return int(f"{annee % 100:02d}{1000 + idx + 1:05d}")


def fetch_invoices(type_inv: str, mois: int, annee: int) -> list:
    _, last_day = monthrange(annee, mois)
    date_min = f"{annee}-{mois:02d}-01"
    date_max = f"{annee}-{mois:02d}-{last_day:02d}"
    result = (
        supabase.table("invoices")
        .select("*")
        .eq("type", type_inv)
        .gte("invoice_date", date_min)
        .lte("invoice_date", date_max)
        .order("invoice_date")
        .execute()
    )
    return result.data or []


def fill_saisie(ws, invoices: list, mois: int, annee: int, is_achat: bool):
    ws["B4"] = mois
    ws["B5"] = annee

    max_row = max(ws.max_row, 10 + len(invoices) + 5)
    for r in range(10, max_row + 1):
        for c in range(2, 21):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and not str(cell.value).startswith("="):
                cell.value = None

    tva_col = 17 if is_achat else 10

    for i, inv in enumerate(invoices):
        row = 10 + i
        ttc = inv.get("amount_ttc") or 0
        ht = inv.get("amount_ht")
        tva = round(ttc - ht, 2) if ht is not None else None

        ws.cell(row=row, column=2).value = get_day(inv)
        ws.cell(row=row, column=4).value = ttc
        ws.cell(row=row, column=5).value = make_piece_number(annee, i)
        ws.cell(row=row, column=6).value = build_libelle(inv)
        ws.cell(row=row, column=7).value = ht
        ws.cell(row=row, column=tva_col).value = tva


def generate_excel(type_inv: str, mois: int, annee: int) -> tuple[io.BytesIO, int]:
    is_achat = type_inv == "achat"
    template_name = "achats_template.xlsx" if is_achat else "ventes_template.xlsx"
    template_path = os.path.join(TEMPLATES_DIR, template_name)

    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Template '{template_name}' introuvable dans {TEMPLATES_DIR}/. "
            "Ajoute les fichiers templates dans ce dossier."
        )

    invoices = fetch_invoices(type_inv, mois, annee)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["SAISIE"]
    fill_saisie(ws, invoices, mois, annee, is_achat)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, len(invoices)


# ─── Routes ───────────────────────────────────────────

@router.get("/export")
async def export_page(request: Request):
    guard = require_auth(request)
    if guard:
        return guard

    now = date.today()
    return templates.TemplateResponse("export.html", {
        "request": request,
        "user_email": get_user_email(request),
        "mois_list": MOIS_FR,
        "current_month": now.month,
        "current_year": now.year,
        "years": list(range(now.year, now.year - 5, -1)),
    })


@router.get("/export/preview")
async def export_preview(
    request: Request,          # ← ajouté
    mois: int = Query(...),
    annee: int = Query(...),
    type_export: str = Query(...),
):
    guard = require_auth(request)  # ← ajouté
    if guard:
        return guard

    invoices = fetch_invoices(type_export, mois, annee)
    total_ttc = sum(inv.get("amount_ttc") or 0 for inv in invoices)
    return {
        "count": len(invoices),
        "total_ttc": round(total_ttc, 2),
    }


@router.get("/export/download")
async def export_download(
    request: Request,
    mois: int = Query(...),
    annee: int = Query(...),
    type_export: str = Query(...),
):
    guard = require_auth(request)
    if guard:
        return guard

    try:
        output, count = generate_excel(type_export, mois, annee)
    except FileNotFoundError as e:
        return HTMLResponse(
            f"<p style='color:red;font-family:sans-serif'><strong>Erreur :</strong> {e}</p>",
            status_code=500,
        )

    type_label = "Achats" if type_export == "achat" else "Ventes"
    filename = f"{type_label}_{mois:02d}_{annee}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )